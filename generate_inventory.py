# -*- coding: utf-8 -*-
"""扫描软件安装包目录，生成/更新清单 Excel 表格。
第一次：全部生成。后续：累加增量，按文件名去重，已有条目描述自动刷新。
双击 更新软件清单.bat 即可执行。
"""
import os, re, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------- 配置 ----------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DIRS = {
    "办公常用": os.path.join(SCRIPT_DIR, "办公常用"),
    "开发常用": os.path.join(SCRIPT_DIR, "开发常用"),
}
OUTPUT = os.path.join(SCRIPT_DIR, "软件清单.xlsx")
HEADERS = ["ID", "来源", "文件名", "描述"]

# ---------- 描述映射 ----------
DESC_MAP = [
    # 办公常用（长关键词在前，避免短关键词误匹配）
    ("office20242016", "Office 2024 与 2016 合集/安装器"),
    ("Office2016", "Microsoft Office 2016 办公套件"),
    ("Office2024", "Microsoft Office 2024 办公套件"),
    ("360zip", "360压缩，压缩解压工具"),
    ("7z", "7-Zip 高压缩比解压工具"),
    ("AweSun", "AweSun 远程桌面控制软件"),
    ("BaiduNetdisk", "百度网盘 Windows 客户端"),
    ("Bypass", "百度网盘不限速下载辅助工具"),
    ("CAJViewer", "CAJ 论文/学术文献阅读器"),
    ("cmder", "Cmder 终端模拟器（类 Linux 控制台）"),
    ("cv_x64", "CV 工具/计算机视觉相关软件"),
    ("dgsetup", "DiskGenius 磁盘分区与数据恢复工具"),
    ("Evernote", "印象笔记，云笔记管理工具"),
    ("Everything", "Everything 极速文件搜索工具"),
    ("listary", "Listary 文件搜索与快速启动工具"),
    ("notepad", "Notepad++ 轻量级代码/文本编辑器"),
    ("Quark", "夸克浏览器 PC 版"),
    ("sogou", "搜狗拼音输入法"),
    ("ToDesk", "ToDesk 远程控制软件"),
    ("WeChat", "微信 Windows 桌面版"),
    ("WorkBuddy", "WorkBuddy AI 编程助手客户端"),
    ("WPS", "WPS Office 办公套件"),
    ("Win10", "Windows 10 企业版系统镜像文件"),
    ("虚拟机", "虚拟机软件（VMware/VirtualBox 等）"),
    ("速鹰", "速鹰下载工具/下载器"),
    # 开发常用
    ("Codex", "OpenAI Codex AI 代码生成工具"),
    ("Git-", "Git 分布式版本控制系统"),
    ("node-", "Node.js JavaScript 运行时环境"),
    ("Ollama", "Ollama 本地大模型运行框架"),
    ("VSCode", "Visual Studio Code 代码编辑器"),
    ("anconda", "Anaconda3 Python 数据科学发行版"),
    ("jmeter", "Apache JMeter 性能/压力测试工具"),
    ("Matlab", "MATLAB 2023b 数值计算与仿真软件"),
    ("postman", "Postman API 开发与测试工具"),
    ("pycharms", "JetBrains PyCharm Python IDE"),
    ("suying", "速鹰下载工具"),
    ("Xmind", "XMind 2024 思维导图工具"),
]


def get_description(name: str) -> str:
    """根据文件名关键词匹配描述。"""
    name_lower = name.lower()
    for keyword, desc in DESC_MAP:
        if keyword.lower() in name_lower:
            return desc
    base = re.sub(r"[-_]?(v?[\d.]+[\d]|x64|x86|win\d*|setup|installer)", "", name_lower, flags=re.I)
    base = base.strip("-_ .")
    if base:
        return f"{base} 软件安装包"
    return "未知软件"


def scan_dirs():
    """扫描所有目录，返回 [来源, 文件名] 列表。"""
    items = []
    for source, path in DIRS.items():
        if not os.path.exists(path):
            print(f"  [警告] 目录不存在: {path}")
            continue
        for entry in os.listdir(path):
            full = os.path.join(path, entry)
            if os.path.isfile(full) or os.path.isdir(full):
                items.append((source, entry))
    return items


def load_existing(path):
    """加载已有 Excel，返回 {文件名: (id, 来源, 描述)} 字典。"""
    if not os.path.exists(path):
        return {}, 1
    wb = load_workbook(path)
    ws = wb.active
    existing = {}
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rid = int(row[0]) if row[0] else 0
        src = str(row[1]) if row[1] else ""
        name = str(row[2]) if row[2] else ""
        desc = str(row[3]) if row[3] else ""
        if name:
            existing[name] = (rid, src, desc)
        max_id = max(max_id, rid)
    wb.close()
    return existing, max_id + 1


def style_sheet(ws):
    """统一表格样式。"""
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="微软雅黑", size=10)
            if cell.column == 1:
                cell.alignment = Alignment(horizontal="center")
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 40


def main():
    print("=" * 60)
    print("  软件清单生成工具")
    print("=" * 60)
    print()
    print("扫描目录...")
    new_items = scan_dirs()
    print(f"  共发现 {len(new_items)} 个项目")

    print("检查已有清单...")
    existing, next_id = load_existing(OUTPUT)
    print(f"  已有 {len(existing)} 条记录，下一 ID：{next_id}")

    # 旧记录永不删除，先全部保留
    merged = existing.copy()
    new_count = 0
    update_count = 0
    kept_count = 0  # 磁盘上已不存在的旧记录，仍保留在表中

    for source, name in new_items:
        new_desc = get_description(name)
        if name in existing:
            old_id, old_src, old_desc = existing[name]
            # 描述优先用最新匹配的，来源用磁盘上当前的
            merged[name] = (old_id, source, new_desc)
            if old_desc != new_desc:
                update_count += 1
            del existing[name]  # 标记为"仍存在"
        else:
            merged[name] = (next_id, source, new_desc)
            next_id += 1
            new_count += 1

    # existing 里剩下的就是磁盘上已不存在的旧记录，保留不动
    kept_count = len(existing)

    print(f"  新增 {new_count} 条，描述更新 {update_count} 条，保留历史记录 {kept_count} 条")

    wb = Workbook()
    ws = wb.active
    ws.title = "软件清单"

    sorted_items = sorted(merged.values(), key=lambda x: x[0])
    for row_idx, (rid, src, desc) in enumerate(sorted_items, 2):
        ws.cell(row=row_idx, column=1, value=rid)
        ws.cell(row=row_idx, column=2, value=src)
        for k, v in merged.items():
            if v[0] == rid:
                ws.cell(row=row_idx, column=3, value=k)
                break
        ws.cell(row=row_idx, column=4, value=desc)

    style_sheet(ws)
    wb.save(OUTPUT)

    print()
    print(f"[完成] 已保存: {OUTPUT}")
    print(f"  总计 {len(merged)} 条记录（新增 {new_count}，继承 {len(merged) - new_count}）")


if __name__ == "__main__":
    main()
